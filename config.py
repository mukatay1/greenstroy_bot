from openpyxl.styles import PatternFill, Border, Side


DEBUG = False

months_russian = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь"
}

# supervisors = [
#     "Кексель Кристина",
#     "Тайбупенова Шолпан"
# ]

supervisors = [
    "Руководитель - 1",
    "Руководитель - 2"
]

weekdays = [
    'Понедельник',
    'Вторник',
    'Среда',
    'Четверг',
    'Пятница',
    'Суббота',
    'Воскресенье'
]

red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
black_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)
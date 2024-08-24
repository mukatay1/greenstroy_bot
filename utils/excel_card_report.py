import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell

from config import black_border
from database.models.City import City


def excel_card_report(data, columns, rus_month, city: City):
    df = pd.DataFrame(data, columns=columns)

    file_path = f'Табель за {rus_month} - г.{city.value}.xlsx'
    df.to_excel(file_path, index=False, engine='openpyxl')

    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = f'Табель за {rus_month}'

    ws.insert_rows(1)
    last_column_letter = get_column_letter(len(columns))
    ws.merge_cells(f'A1:{last_column_letter}1')
    ws['A1'] = f'Табель за {rus_month} - г.{city.value}'
    ws['A1'].font = Font(size=16, bold=True)
    # ws['A1'].alignment = Alignment(horizontal='center')

    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    for row in ws.iter_rows(min_row=1, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.border = black_border

    merge_cells_in_triples(ws, 'A', 3)
    merge_cells_in_triples(ws, 'B', 3)
    merge_cells_in_triples(ws, last_column_letter, 3, True)
    wb.save(file_path)
    return file_path


def merge_cells_in_triples(sheet, column_letter, start_row, last = False):
    max_row = sheet.max_row
    for row in range(start_row, max_row, 3):
        cell1 = sheet[f'{column_letter}{row}']
        cell2 = sheet[f'{column_letter}{row + 1}']
        cell3 = sheet[f'{column_letter}{row + 2}']
        sheet.merge_cells(f'{column_letter}{row}:{column_letter}{row + 2}')
        if last:
            cell1.value = f'{cell3.value}'
            cell2.value = None
            cell3.value = None
        else:
            cell1.value = f'{cell1.value}'
            cell2.value = None
            cell3.value = None
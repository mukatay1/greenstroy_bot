import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font

from database.models.City import City
from utils.adjust_column_width import adjust_column_width
from utils.apply_border import apply_border


def excel_report(data, report_date, city: City):
    df = pd.DataFrame(data)
    report_file = f"Отчёт_{report_date}_{city.value}.xlsx"
    df.to_excel(report_file, index=False, engine='openpyxl')

    wb = load_workbook(report_file)
    ws = wb.active
    ws.title = "Отчет"

    ws.insert_rows(1)
    ws.merge_cells('A1:L1')
    ws['A1'] = f'Отчет сотрудников за {report_date} - г.{city.value}'
    ws['A1'].font = Font(size=16, bold=True)

    adjust_column_width(ws, df)
    apply_border(ws, len(df.columns))

    wb.save(report_file)
    return report_file

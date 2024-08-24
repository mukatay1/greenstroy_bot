import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from config import red_fill, black_border, green_fill
from database.models.City import City
from utils.adjust_column_width import adjust_column_width


def late_excel_report(data: list[dict], month_name: str, city: City) -> str:
    df = pd.DataFrame(sorted(data, key=lambda x: x["Количество опозданий"], reverse=True))

    report_file = f"Отчет_по_опозданиям_за_{month_name}_{city.value}.xlsx"
    df.to_excel(report_file, index=False, engine='openpyxl')

    wb = load_workbook(report_file)
    ws = wb.active
    ws.title = "Отчет"

    ws.insert_rows(1)
    ws.merge_cells('A1:E1')
    ws['A1'] = f'Отчет по опозданиям сотрудников за {month_name} - г.{city.value}'
    ws['A1'].font = Font(size=16, bold=True)

    adjust_column_width(ws, df)

    for row in ws.iter_rows(min_row=3, max_col=5):
        try:
            late_count = int(row[3].value)
        except (ValueError, TypeError):
            late_count = 0

        if late_count > 3:
            for cell in row:
                cell.fill = red_fill
                cell.border = black_border
        else:
            for cell in row:
                cell.fill = green_fill
                cell.border = black_border

    wb.save(report_file)
    return report_file
